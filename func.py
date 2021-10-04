import os
from zipfile import ZipFile
import re
import openpyxl

# import nltk #for those who haven't installed nltk and subordinate libraries
# nltk.download('punkt')
# nltk.download('wordnet')

from nltk.stem import WordNetLemmatizer, PorterStemmer
from nltk.tokenize import sent_tokenize

def analyze_txt(text_file): #should be in format "title.txt"
    rawf = open(text_file, "r")
    raw_txt = rawf.read() #string data of whole text
    raw_txt = raw_txt.lower() #all capital letters to small letters
    
    print("Tokenizing into sentences and words...")
    
    
    # ---------- Preprocess the raw text by tokenizing, lemmatizing, and stemming ----------
    
    
    sent_list = sent_tokenize(raw_txt) #tokenize by sentence
    
    word_list = [] # list of raw word divided by each
    for sent in sent_list:
        word_list.append(list(filter(lambda x: x !="",re.split("\W", sent)))) #sentences in the list are divided into words, and then empty srings are deleted
    
    print("Normalizing the txt...")
    
    lm = WordNetLemmatizer()
    lm_txt = [] # list of words lemmatized from raw_txt
    for sent in word_list:
        lm_txt.append([lm.lemmatize(word, pos="v") for word in sent]) #lemmatized text
    
    ps = PorterStemmer()
    st_txt = [] # list of words stemmed from lm_txt
    for sent in lm_txt:
        st_txt.append([ps.stem(word = word) for word in sent]) #stemmed text
    
    print("Counting the frequencies...")
    
    
    # ---------- Count frequencies of each word and saves in the dictionary 'freq' ----------
    
    
    wb = openpyxl.Workbook() #new workbook
    ws = wb.active

    freq = {} #dictionary that contains words and corresponding frequencies
    sent_count = 0 #index of sentences(first-dimension element) of lem_txt
    stem_count = 0 #index of stems(second-dimension element) of lm_txt
    st_txt_len = len(st_txt) # the total number of words
    for sent in st_txt:
        for stem in sent:
            if freq.get(stem) == None: #if the stem isn't initialized yet, initialize it by 0.
                freq[stem] = {lm_txt[sent_count][stem_count] : 0}
            elif freq[stem].get(lm_txt[sent_count][stem_count]) == None: #if the word isn't initialized yet, initialize it by 0.
                freq[stem][lm_txt[sent_count][stem_count]] = 0

            freq[stem][lm_txt[sent_count][stem_count]] += 1 #new detection, plus one for frequency
            stem_count += 1

        sent_count += 1 # the total number of sentences counted
        stem_count = 0 # reset after counting words from each stem
        
        if sent_count % 500 == 0:
            print(f"Counted {sent_count} out of {st_txt_len}... ({(sent_count/st_txt_len)*100:.0f}%)")

            
    # ---------- Save the result as an exel file(.xlsx) ----------
    
    
    print("Saving as an exel file...")

    row = 1 #index of freq
    word_len = 0
    for stem, words in freq.items():
        word_len += len(freq[stem])
    
    for stem, words in freq.items():
        ws.cell(row = row, column = 1).value = stem #the first column with stems
        ws.merge_cells(start_row = row, start_column = 1, end_row = row + len(freq[stem]) - 1, end_column =  1)
        for word, frequency in freq[stem].items():
            ws.cell(row = row, column = 2).value = word #the second column with words, multiple if needed
            ws.cell(row = row, column = 3).value = frequency #the third column with corresponding frequencies    
            row += 1
            
            if row % 500 == 0:
                print(f"Saved {row} out of {word_len}... ({(row/word_len)*100 :.0f}%)")

    wb.save(filename = f"result_{text_file[:-4]}.xlsx") #save the result worksheet
    wb.close()
    rawf.close()
    
    print("Saved as \"result_{}.xlsx\"".format(text_file[:-4]))
    
    return 0





def analyze_epub(file_name):
    print("Processing the epub file...")
    
    os.rename('%s' % file_name, file_name[:-4] + "zip") # convert from a epub file to a zip file
    with ZipFile(file_name[:-4] + "zip") as zip:
        zip.extractall(file_name[:-4]) # unzip the zip file

    dir = './%s/OEBPS/' % file_name[:-4]

    def is_html(file_name): # determines whether the input file is html
        if file_name[-4:] == 'html':
            return True
        else:
            return False

    file_list = list(filter(is_html, os.listdir('./%s/OEBPS' % file_name[:-4]))) # contains the name of html files stored in the epub
    file_list.sort()

    for file in file_list:
        os.rename(dir + file, dir + file[:-4]+'txt') # convert html files into txt files
        file_list[file_list.index(file)] = file[:-4]+'txt' # updating elements in file_list from html extension to txt

    file_list = list(map(lambda x: dir + x, file_list)) # add directory address to be relative adress

    # html to plain text
    for file in file_list:
        with open(file, 'r') as target:
            orig_text = target.read() # store the original text which is in html format

        with open(file, 'w') as target:
            orig_text = target.write(re.sub("<[^<>]*>","",orig_text)) # distinguish text in the html and rewrite the files with them

    destination = re.sub(dir, "", file_list[0]) # strip the directory address off
    with open(destination, "a+") as merged: # mutiple text files are merged into one
        for file in file_list:
            with open(file) as f:
                merged.write(f.read())

    with os.popen("rm -rf %s" % file_name[:-4]) as p: # delete the unzipped epub file
        pass

    analyze_txt(destination)





def analyze_pdf(pdf_file):
    
    return 0