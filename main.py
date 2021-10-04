import re
import openpyxl

# import nltk #for those who haven't installed nltk and subordinate libraries
# nltk.download('punkt')
# nltk.download('wordnet')

from nltk.stem import WordNetLemmatizer, PorterStemmer
from nltk.tokenize import sent_tokenize

def analyze(text_file): #should be in format "title.txt"
    rawf = open(text_file, "r")
    raw_txt = rawf.read() #string data of whole text
    raw_txt = raw_txt.lower() #all capital letters to small letters

    sent_list = sent_tokenize(raw_txt) #tokenize by sentence
    
    word_list = []
    for sent in sent_list:
        word_list.append(list(filter(lambda x: x !="",re.split("\W", sent)))) #sentences in the list are divided into words, and then empty srings are deleted
    
    lm = WordNetLemmatizer()
    lm_txt = []
    for sent in word_list:
        lm_txt.append([lm.lemmatize(word, pos="v") for word in sent]) #lemmatized text
    
    ps = PorterStemmer()
    st_txt = []
    for sent in lm_txt:
        st_txt.append([ps.stem(word = word) for word in sent]) #stemmed text

    result = open("st.txt", "w") #save st_txt as txt file#
    for s in st_txt:#
        for w in s:#
            result.write(w + "\n")#
    result.write(str(st_txt) + "\n")
    result.close()#
    
    print(st_txt)#
    
    wb = openpyxl.Workbook() #new workbook
    ws = wb.active

    freq = {} #dictionary that contains words and corresponding frequencies
    sent_count = 0 #index of sentences(first-dimension element) of lem_txt
    stem_count = 0 #index of stems(second-dimension element) of lm_txt
    for sent in st_txt:
        for stem in sent:
            if freq.get(stem) == None: #if the stem isn't initialized yet, initialize it by 0.
                freq[stem] = {lm_txt[sent_count][stem_count] : 0}
            elif freq[stem].get(lm_txt[sent_count][stem_count]) == None: #if the word isn't initialized yet, initialize it by 0.
                freq[stem][lm_txt[sent_count][stem_count]] = 0

            freq[stem][lm_txt[sent_count][stem_count]] += 1 #new detection, plus one for frequency
            stem_count += 1

        sent_count += 1
        stem_count = 0

    row = 1 #index of freq
    for stem, words in freq.items():
        ws.cell(row = row, column = 1).value = stem #the first column with stems
        ws.merge_cells(start_row = row, start_column = 1, end_row = row + len(freq[stem]) - 1, end_column =  1)
        for word, frequency in freq[stem].items():
            ws.cell(row = row, column = 2).value = word #the second column with words, multiple if needed
            ws.cell(row = row, column = 3).value = frequency #the third column with corresponding frequencies    
            row += 1
            print(stem, word, frequency, row)
            
        # ws.merge_cells(start_row = row - len(freq[stem]), start_column = 2, end_row = row - 1, end_column =  2)

    wb.save(filename = f"result_{text_file[:-4]}.xlsx") #save the result worksheet
    wb.close()
    rawf.close()

    return 0

if __name__ == "__main__":
    analyze("The_Great_Gatsby.txt")